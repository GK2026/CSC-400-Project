from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import requests
import csv
import json
from io import StringIO, BytesIO
from math import sqrt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.db.session import get_db
from app.models.gapminder_data import GapminderData
from app.models.exercise import Exercise
from app.models.submission import Submission
from app.schemas.gapminder import DatasetRequest, CorrelationRequest
from app.schemas.exercise import SaveExerciseRequest
from app.schemas.submission import SubmitExerciseRequest, SubmissionFeedbackRequest
from app.core.deps import get_current_user
from app.models.user import User

# router
router = APIRouter()

# github config
OWNER = "open-numbers"
REPO = "ddf--gapminder--systema_globalis"
BRANCH = "master"
TARGET_DIR = "countries-etc-datapoints"


# github helpers
def github_get(url: str):
    import os
    token = os.getenv("GITHUB_TOKEN", "")
    headers = {
        "Accept": "application/vnd.github+json",
        "User-Agent": "GapMinders-App"
    }
    if token:
        headers["Authorization"] = f"Bearer {token}"
    return requests.get(url, headers=headers, timeout=30)


# repo tree
def get_target_dir_tree():
    root_tree_url = f"https://api.github.com/repos/{OWNER}/{REPO}/git/trees/{BRANCH}"
    root_resp = github_get(root_tree_url)

    if root_resp.status_code != 200:
        return None, {
            "error": "Failed to read repo root from GitHub API",
            "status_code": root_resp.status_code
        }

    root_items = root_resp.json().get("tree", [])
    target_dir_sha = None

    for item in root_items:
        if item.get("type") == "tree" and item.get("path") == TARGET_DIR:
            target_dir_sha = item.get("sha")
            break

    if not target_dir_sha:
        return None, {"error": f"Directory '{TARGET_DIR}' not found"}

    sub_tree_url = f"https://api.github.com/repos/{OWNER}/{REPO}/git/trees/{target_dir_sha}?recursive=1"
    sub_resp = github_get(sub_tree_url)

    if sub_resp.status_code != 200:
        return None, {
            "error": "Failed to read subtree from GitHub API",
            "status_code": sub_resp.status_code
        }

    return sub_resp.json().get("tree", []), None


def build_full_path(filename: str, tree_items: list[dict]) -> str | None:
    for item in tree_items:
        if item.get("type") == "blob" and item.get("path", "").endswith(filename):
            return f"{TARGET_DIR}/{item['path']}"
    return None


def get_download_url(full_path: str):
    contents_url = f"https://api.github.com/repos/{OWNER}/{REPO}/contents/{full_path}?ref={BRANCH}"
    resp = github_get(contents_url)

    if resp.status_code != 200:
        return None, {
            "error": "Failed to fetch file metadata from GitHub Contents API",
            "status_code": resp.status_code,
            "path": full_path
        }

    data = resp.json()
    return data.get("download_url"), None


# indicator helpers
def infer_indicator_code(filename: str) -> str:
    prefix = "ddf--datapoints--"
    suffix = "--by--geo--time.csv"

    if filename.startswith(prefix) and filename.endswith(suffix):
        return filename[len(prefix):-len(suffix)]

    return filename.replace(".csv", "")


def classify_correlation(r: float) -> str:
    if r > 0:
        return "positive"
    if r < 0:
        return "negative"
    return "none"


def calculate_pearson_from_rows(rows: list[dict]) -> float:
    n = len(rows)

    if n < 2:
        raise HTTPException(
            status_code=400,
            detail="At least 2 valid paired data points are required to calculate correlation."
        )

    x_values = [row["indicator_1_value"] for row in rows]
    y_values = [row["indicator_2_value"] for row in rows]

    sum_x = sum(x_values)
    sum_y = sum(y_values)
    sum_x2 = sum(x * x for x in x_values)
    sum_y2 = sum(y * y for y in y_values)
    sum_xy = sum(x * y for x, y in zip(x_values, y_values))

    numerator = (n * sum_xy) - (sum_x * sum_y)
    denominator_left = (n * sum_x2) - (sum_x * sum_x)
    denominator_right = (n * sum_y2) - (sum_y * sum_y)
    denominator = sqrt(denominator_left * denominator_right)

    if denominator == 0:
        raise HTTPException(
            status_code=400,
            detail="Correlation cannot be calculated because one indicator has no variation."
        )

    return numerator / denominator


def require_teacher(current_user: User):
    if current_user.role not in ("teacher", "instructor"):
        raise HTTPException(status_code=403, detail="Teacher access required.")


def make_csv_response(csv_text: str, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([csv_text]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# xlsx export
def build_dataset_xlsx(
    indicator_1_name: str,
    indicator_2_name: str,
    start_year: int,
    end_year: int,
    countries: list[str],
    rows: list[dict],
    pearson_r: float | None,
    relationship_label: str | None,
    points_used: int,
) -> bytes:
    from openpyxl.styles import Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dataset"

    is_over_time = start_year != end_year

    # styles
    purple       = "7C3AED"
    light_purple = "F3E8FF"
    mid_purple   = "EDE9FE"
    white        = "FFFFFF"
    dark_text    = "1F1F1F"
    muted_text   = "6B7280"

    header_font   = Font(bold=True, color=white, size=11)
    header_fill   = PatternFill("solid", fgColor=purple)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

    title_font    = Font(bold=True, color=purple, size=14)
    label_font    = Font(bold=True, color=dark_text, size=10)
    value_font    = Font(color=dark_text, size=10)
    muted_font    = Font(italic=True, color=muted_text, size=9)

    label_fill    = PatternFill("solid", fgColor=light_purple)
    alt_fill      = PatternFill("solid", fgColor=mid_purple)
    plain_fill    = PatternFill("solid", fgColor=white)

    thin_border   = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    thick_bottom  = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="medium", color=purple),
    )

    def set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None, number_format=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font:        cell.font = font
        if fill:        cell.fill = fill
        if alignment:   cell.alignment = alignment
        if border:      cell.border = border
        if number_format: cell.number_format = number_format
        return cell

    r = 1

    # title row
    ws.row_dimensions[r].height = 28
    set_cell(ws, r, 1, "Correlation Assistant — Student Worksheet",
             font=title_font,
             alignment=Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    # blank gap
    ws.row_dimensions[r].height = 6
    r += 1

    # info block
    info_rows = [("Indicator X", indicator_1_name), ("Indicator Y", indicator_2_name)]
    if is_over_time:
        info_rows.append(("Country", countries[0] if countries else ""))
        info_rows.append(("Year Range", f"{start_year} to {end_year}"))
    else:
        info_rows.append(("Year", str(start_year)))
        info_rows.append(("Countries", ", ".join(countries) if countries else ""))
    info_rows.append(("Data Points", str(points_used)))

    for lbl, val in info_rows:
        ws.row_dimensions[r].height = 18
        set_cell(ws, r, 1, lbl, font=label_font, fill=label_fill,
                 alignment=Alignment(horizontal="left", vertical="center"),
                 border=thin_border)
        set_cell(ws, r, 2, val, font=value_font, fill=plain_fill,
                 alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                 border=thin_border)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1

    r += 1  # gap

    # instructions row
    ws.row_dimensions[r].height = 36
    instr = ("Use the X and Y values below to calculate your correlation. "
             "Record your work in the Correlation Guide.")
    set_cell(ws, r, 1, instr,
             font=muted_font,
             alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    # column headers
    col1_label = "Year" if is_over_time else "Country"
    headers = [col1_label, f"X  —  {indicator_1_name}", f"Y  —  {indicator_2_name}"]
    ws.row_dimensions[r].height = 36
    for ci, h in enumerate(headers, start=1):
        set_cell(ws, r, ci, h,
                 font=header_font,
                 fill=header_fill,
                 alignment=header_align,
                 border=thick_bottom)
    header_data_row = r
    r += 1

    # data rows
    for i, row in enumerate(rows):
        ws.row_dimensions[r].height = 18
        label = row.get("year", "") if is_over_time else (row.get("country") or row.get("country_code", ""))
        fill = alt_fill if i % 2 == 1 else plain_fill

        set_cell(ws, r, 1, label,      font=value_font, fill=fill,
                 alignment=Alignment(horizontal="center", vertical="center"), border=thin_border)
        set_cell(ws, r, 2, row["indicator_1_value"], font=value_font, fill=fill,
                 alignment=Alignment(horizontal="center", vertical="center"), border=thin_border,
                 number_format="0.000")
        set_cell(ws, r, 3, row["indicator_2_value"], font=value_font, fill=fill,
                 alignment=Alignment(horizontal="center", vertical="center"), border=thin_border,
                 number_format="0.000")

        r += 1

    r += 1

    # result area
    r += 1
    set_cell(ws, r, 1, "My Results",
             font=Font(bold=True, color=purple, size=11),
             alignment=Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    # instructions row
    ws.row_dimensions[r].height = 40
    instr2 = (
        "In Excel or Google Sheets, use =CORREL(X_range, Y_range) on the data above to get your correlation value. "
        "Enter the result in the cell to the right."
    )
    set_cell(ws, r, 1, instr2,
             font=muted_font,
             alignment=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    r += 1

    # Correlation entry row
    ws.row_dimensions[r].height = 24
    set_cell(ws, r, 1, "Correlation  =", font=label_font, fill=label_fill,
             alignment=Alignment(horizontal="left", vertical="center"), border=thin_border)
    set_cell(ws, r, 2, "", fill=plain_fill, border=thin_border)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 1

    # column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = max(28, min(len(indicator_1_name) + 6, 48))
    ws.column_dimensions["C"].width = max(28, min(len(indicator_2_name) + 6, 48))


    # no freeze panes - fully scrollable

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# csv export
def build_dataset_csv_text(
    indicator_1_code: str,
    indicator_1_name: str,
    indicator_2_code: str,
    indicator_2_name: str,
    start_year: int,
    end_year: int,
    countries: list[str],
    rows: list[dict],
    pearson_r: float | None,
    relationship_label: str | None,
    points_used: int,
    points_skipped: int
) -> str:
    output = StringIO()
    writer = csv.writer(output)


    is_over_time = start_year != end_year


    writer.writerow(["Correlation Assistant Export"])
    writer.writerow([])
    writer.writerow(["Indicator 1", indicator_1_name])
    writer.writerow(["Indicator 2", indicator_2_name])
    if is_over_time:
        writer.writerow(["Country", countries[0] if countries else ""])
        writer.writerow(["Years", f"{start_year} - {end_year}"])
    else:
        writer.writerow(["Year", start_year])
        writer.writerow(["Countries", ", ".join(countries) if countries else ""])
    writer.writerow(["Data Points Used", points_used])
    writer.writerow([])


    if is_over_time:
        writer.writerow(["Year", indicator_1_name, indicator_2_name])
        for row in rows:
            writer.writerow([row["year"], row["indicator_1_value"], row["indicator_2_value"]])
    else:
        writer.writerow(["Country", indicator_1_name, indicator_2_name])
        for row in rows:
            country_label = row.get("country") or row.get("country_code", "")
            writer.writerow([country_label, row["indicator_1_value"], row["indicator_2_value"]])

    return output.getvalue()


# list datasets
@router.get("/datasets")
def list_datasets(limit: int = 50):
    tree_items, err = get_target_dir_tree()
    if err:
        return err

    files = [
        item["path"]
        for item in tree_items
        if item.get("type") == "blob" and item.get("path", "").endswith(".csv")
    ]

    files.sort()
    return {
        "count": len(files),
        "datasets": files[:limit]
    }


# import single
@router.post("/import")
def import_gapminder_data(
    filename: str = Query(..., description="Exact CSV filename from /gapminder/datasets"),
    db: Session = Depends(get_db)
):
    tree_items, err = get_target_dir_tree()
    if err:
        return err

    full_path = build_full_path(filename, tree_items)
    if not full_path:
        return {"error": "Dataset file not found in Gapminder repo", "filename": filename}

    download_url, err = get_download_url(full_path)
    if err:
        return err

    if not download_url:
        return {"error": "No download_url returned by GitHub API", "path": full_path}

    resp = requests.get(download_url, timeout=60)
    if resp.status_code != 200:
        return {
            "error": "Failed to download dataset file",
            "status_code": resp.status_code,
            "path": full_path
        }

    reader = csv.DictReader(StringIO(resp.text))
    indicator_code = infer_indicator_code(filename)

    saved = 0
    skipped = 0
    sample_columns = reader.fieldnames

    try:
        for row in reader:
            country_code = row.get("geo")
            year_raw = row.get("time")
            value_raw = row.get(indicator_code)

            if not country_code or not year_raw or value_raw in (None, ""):
                skipped += 1
                continue

            try:
                year = int(year_raw)
                value = float(value_raw)
            except ValueError:
                skipped += 1
                continue

            exists = db.query(GapminderData).filter(
                GapminderData.country_code == country_code,
                GapminderData.indicator_code == indicator_code,
                GapminderData.year == year
            ).first()

            if exists:
                skipped += 1
                continue


            country_name = row.get("geo.name") or row.get("country_name") or country_code

            record = GapminderData(
                country=country_name,
                country_code=country_code,
                indicator=" ".join(w.capitalize() if len(w) > 3 else w for w in indicator_code.replace("_", " ").split()),
                indicator_code=indicator_code,
                year=year,
                value=value
            )

            db.add(record)
            saved += 1

        db.commit()

        return {
            "message": "Gapminder dataset imported through GitHub API",
            "filename": filename,
            "path_used": full_path,
            "indicator_code": indicator_code,
            "rows_saved": saved,
            "rows_skipped": skipped,
            "columns": sample_columns
        }

    except Exception as e:
        db.rollback()
        return {
            "error": str(e),
            "filename": filename,
            "path_used": full_path,
            "indicator_code": indicator_code,
            "columns": sample_columns
        }


# import all
@router.post("/import-all")
def import_all_gapminder_data(db: Session = Depends(get_db)):
    from sqlalchemy import text

    tree_items, err = get_target_dir_tree()
    if err:
        return err

    files = [
        item["path"]
        for item in tree_items
        if item.get("type") == "blob" and item.get("path", "").endswith(".csv")
    ]

    # get already-imported indicator codes so we can skip them entirely
    existing_codes = {
        row[0] for row in db.execute(
            text("SELECT DISTINCT indicator_code FROM gapminder_data")
        ).fetchall()
    }

    results = []
    total_saved = 0
    total_skipped = 0

    for filename in files:
        indicator_code = infer_indicator_code(filename)

        if indicator_code in existing_codes:
            results.append({"filename": filename, "rows_saved": 0, "rows_skipped": 0, "note": "already imported"})
            continue

        full_path = build_full_path(filename, tree_items)
        if not full_path:
            results.append({"filename": filename, "error": "File not found in tree"})
            continue

        download_url, err = get_download_url(full_path)
        if err or not download_url:
            results.append({"filename": filename, "error": "Could not get download URL"})
            continue

        try:
            resp = requests.get(download_url, timeout=60)
            if resp.status_code != 200:
                results.append({"filename": filename, "error": f"HTTP {resp.status_code}"})
                continue

            reader = csv.DictReader(StringIO(resp.text))
            saved = 0
            skipped = 0
            batch = []

            for row in reader:
                country_code = row.get("geo")
                year_raw = row.get("time")
                value_raw = row.get(indicator_code)

                if not country_code or not year_raw or value_raw in (None, ""):
                    skipped += 1
                    continue

                try:
                    year = int(year_raw)
                    value = float(value_raw)
                except ValueError:
                    skipped += 1
                    continue

                country_name = row.get("geo.name") or row.get("country_name") or country_code
                indicator_name = " ".join(w.capitalize() if len(w) > 3 else w for w in indicator_code.replace("_", " ").split())

                batch.append({
                    "country": country_name,
                    "country_code": country_code,
                    "indicator": indicator_name,
                    "indicator_code": indicator_code,
                    "year": year,
                    "value": value
                })
                saved += 1

            if batch:
                db.bulk_insert_mappings(GapminderData, batch)
                db.commit()
                existing_codes.add(indicator_code)

            total_saved += saved
            total_skipped += skipped
            results.append({"filename": filename, "rows_saved": saved, "rows_skipped": skipped})

        except Exception as e:
            db.rollback()
            results.append({"filename": filename, "error": str(e)})

        except Exception as e:
            db.rollback()
            results.append({"filename": filename, "error": str(e)})

    return {
        "message": "Bulk import complete",
        "files_processed": len(files),
        "total_rows_saved": total_saved,
        "total_rows_skipped": total_skipped,
        "details": results
    }


# data status
@router.get("/data-status")
def get_data_status(db: Session = Depends(get_db)):
    count = db.query(GapminderData).count()
    indicator_count = db.query(GapminderData.indicator_code).distinct().count()
    return {
        "row_count": count,
        "indicator_count": indicator_count,
        "has_data": count > 0
    }


# indicators
_indicators_cache: list | None = None

@router.get("/indicators")
def search_indicators(
    search: str = Query("", description="Search indicator codes or names"),
    limit: int = Query(200, ge=1, le=500),
    db: Session = Depends(get_db)
):
    global _indicators_cache

    if _indicators_cache is None:
        rows = (
            db.query(GapminderData.indicator_code, GapminderData.indicator)
            .distinct()
            .order_by(GapminderData.indicator_code.asc())
            .all()
        )
        _indicators_cache = [
            {"indicator_code": r.indicator_code, "indicator_name": r.indicator}
            for r in rows
        ]

    results = _indicators_cache

    if search.strip():
        term = search.strip().lower()
        results = [
            r for r in results
            if term in r["indicator_code"].lower() or term in r["indicator_name"].lower()
        ]

    return results[:limit]


# generate dataset
@router.post("/generate-dataset")
def generate_dataset(
    payload: DatasetRequest,
    db: Session = Depends(get_db)
):
    if payload.start_year > payload.end_year:
        raise HTTPException(status_code=400, detail="start_year cannot be greater than end_year.")

    country_codes = [code.strip().lower() for code in payload.countries if code.strip()]

    indicator_1_rows_query = db.query(GapminderData).filter(
        GapminderData.indicator_code == payload.indicator_1,
        GapminderData.year >= payload.start_year,
        GapminderData.year <= payload.end_year
    )

    indicator_2_rows_query = db.query(GapminderData).filter(
        GapminderData.indicator_code == payload.indicator_2,
        GapminderData.year >= payload.start_year,
        GapminderData.year <= payload.end_year
    )

    if country_codes:
        indicator_1_rows_query = indicator_1_rows_query.filter(GapminderData.country_code.in_(country_codes))
        indicator_2_rows_query = indicator_2_rows_query.filter(GapminderData.country_code.in_(country_codes))

    indicator_1_rows = indicator_1_rows_query.all()
    indicator_2_rows = indicator_2_rows_query.all()

    indicator_1_map = {}
    for row in indicator_1_rows:
        key = (row.country_code, row.year)
        indicator_1_map[key] = row

    indicator_2_map = {}
    for row in indicator_2_rows:
        key = (row.country_code, row.year)
        indicator_2_map[key] = row

    matching_keys = sorted(set(indicator_1_map.keys()) & set(indicator_2_map.keys()))

    merged_rows = []
    skipped_rows = 0

    for key in matching_keys:
        row1 = indicator_1_map[key]
        row2 = indicator_2_map[key]

        if row1.value is None or row2.value is None:
            skipped_rows += 1
            continue

        try:
            value_1 = float(row1.value)
            value_2 = float(row2.value)
        except (TypeError, ValueError):
            skipped_rows += 1
            continue

        merged_rows.append({
            "country_code": row1.country_code,
            "country": row1.country if row1.country else row1.country_code,
            "year": row1.year,
            "indicator_1_value": value_1,
            "indicator_2_value": value_2
        })

    if not merged_rows:
        raise HTTPException(
            status_code=404,
            detail="No matching clean data was found for those indicators/countries/years."
        )

    indicator_1_name = indicator_1_rows[0].indicator if indicator_1_rows else payload.indicator_1
    indicator_2_name = indicator_2_rows[0].indicator if indicator_2_rows else payload.indicator_2

    return {
        "indicator_1": {
            "code": payload.indicator_1,
            "name": indicator_1_name
        },
        "indicator_2": {
            "code": payload.indicator_2,
            "name": indicator_2_name
        },
        "countries_requested": country_codes,
        "start_year": payload.start_year,
        "end_year": payload.end_year,
        "points_used": len(merged_rows),
        "points_skipped": skipped_rows,
        "rows": merged_rows
    }


# correlation
@router.post("/correlation")
def calculate_correlation(payload: CorrelationRequest):
    rows = [row.model_dump() for row in payload.rows]
    pearson_r = calculate_pearson_from_rows(rows)

    return {
        "pearson_r": round(pearson_r, 6),
        "relationship_label": classify_correlation(pearson_r),
        "points_used": len(rows)
    }


# save exercise
@router.post("/save-exercise")
def save_exercise(
    payload: SaveExerciseRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    rows = [row.model_dump() for row in payload.rows]

    if not rows:
        raise HTTPException(status_code=400, detail="Cannot save an exercise with no rows.")

    pearson_r = payload.pearson_r
    relationship_label = payload.relationship_label

    if pearson_r is None:
        pearson_r = calculate_pearson_from_rows(rows)

    if relationship_label is None:
        relationship_label = classify_correlation(pearson_r)

    exercise = Exercise(
    user_id=current_user.id,

    name=payload.name or "Untitled Exercise",

    indicator_1_code=payload.indicator_1_code,
    indicator_1_name=payload.indicator_1_name,
    indicator_2_code=payload.indicator_2_code,
    indicator_2_name=payload.indicator_2_name,
    countries_json=json.dumps(payload.countries),
    start_year=payload.start_year,
    end_year=payload.end_year,
    merged_rows_json=json.dumps(rows),
    pearson_r=pearson_r,
    relationship_label=relationship_label,
    points_used=payload.points_used if payload.points_used is not None else len(rows),
    points_skipped=payload.points_skipped if payload.points_skipped is not None else 0
    )

    db.add(exercise)
    db.commit()
    db.refresh(exercise)

    return {
        "message": "Exercise saved successfully",
        "exercise_id": exercise.id,
        "pearson_r": round(exercise.pearson_r, 6) if exercise.pearson_r is not None else None,
        "relationship_label": exercise.relationship_label
    }


# my exercises
@router.get("/my-exercises")
def get_my_exercises(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    exercises = (
        db.query(Exercise)
        .filter(Exercise.user_id == current_user.id)
        .order_by(Exercise.created_at.desc())
        .all()
    )

    results = []
    for exercise in exercises:
        results.append({
    "id": exercise.id,
    "name": exercise.name,
    "indicator_1_code": exercise.indicator_1_code,
    "indicator_1_name": exercise.indicator_1_name,
    "indicator_2_code": exercise.indicator_2_code,
    "indicator_2_name": exercise.indicator_2_name,
    "countries": json.loads(exercise.countries_json),
    "start_year": exercise.start_year,
    "end_year": exercise.end_year,
    "pearson_r": exercise.pearson_r,
    "relationship_label": exercise.relationship_label,
    "points_used": exercise.points_used,
    "points_skipped": exercise.points_skipped,
    "created_at": exercise.created_at
    })

    return results


@router.get("/my-exercises/{exercise_id}")
def get_one_exercise(
    exercise_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    exercise = (
        db.query(Exercise)
        .filter(
            Exercise.id == exercise_id,
            Exercise.user_id == current_user.id
        )
        .first()
    )

    if not exercise:
        raise HTTPException(status_code=404, detail="Exercise not found.")

    return {
    "id": exercise.id,
    "name": exercise.name,

    "indicator_1_code": exercise.indicator_1_code,
    "indicator_1_name": exercise.indicator_1_name,
    "indicator_2_code": exercise.indicator_2_code,
    "indicator_2_name": exercise.indicator_2_name,
    "countries": json.loads(exercise.countries_json),
    "start_year": exercise.start_year,
    "end_year": exercise.end_year,
    "rows": json.loads(exercise.merged_rows_json),
    "pearson_r": exercise.pearson_r,
    "relationship_label": exercise.relationship_label,
    "points_used": exercise.points_used,
    "points_skipped": exercise.points_skipped,
    "created_at": exercise.created_at
}


@router.post("/export-csv")
def export_csv_from_payload(
    payload: SaveExerciseRequest,
    current_user: User = Depends(get_current_user)
):
    rows = [row.model_dump() for row in payload.rows]

    if not rows:
        raise HTTPException(status_code=400, detail="Cannot export with no rows.")

    pearson_r = payload.pearson_r
    relationship_label = payload.relationship_label

    if pearson_r is None:
        pearson_r = calculate_pearson_from_rows(rows)

    if relationship_label is None:
        relationship_label = classify_correlation(pearson_r)

    xlsx_bytes = build_dataset_xlsx(
        indicator_1_name=payload.indicator_1_name,
        indicator_2_name=payload.indicator_2_name,
        start_year=payload.start_year,
        end_year=payload.end_year,
        countries=payload.countries,
        rows=rows,
        pearson_r=round(pearson_r, 6),
        relationship_label=relationship_label,
        points_used=payload.points_used if payload.points_used is not None else len(rows),
    )

    filename = f"gapminders_export_{payload.indicator_1_code}_{payload.indicator_2_code}.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.get("/my-exercises/{exercise_id}/export-csv")
def export_saved_exercise_csv(
    exercise_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    exercise = (
        db.query(Exercise)
        .filter(
            Exercise.id == exercise_id,
            Exercise.user_id == current_user.id
        )
        .first()
    )

    if not exercise:
        raise HTTPException(status_code=404, detail="Exercise not found.")

    rows = json.loads(exercise.merged_rows_json)
    countries = json.loads(exercise.countries_json)

    xlsx_bytes = build_dataset_xlsx(
        indicator_1_name=exercise.indicator_1_name,
        indicator_2_name=exercise.indicator_2_name,
        start_year=exercise.start_year,
        end_year=exercise.end_year,
        countries=countries,
        rows=rows,
        pearson_r=round(exercise.pearson_r, 6) if exercise.pearson_r is not None else None,
        relationship_label=exercise.relationship_label,
        points_used=exercise.points_used,
    )

    filename = f"exercise_{exercise.id}_export.xlsx"
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# submit exercise
@router.post("/submit-exercise")
def submit_exercise(
    payload: SubmitExerciseRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    selected_label = payload.student_selected_label.strip().lower()

    if selected_label not in {"positive", "negative", "none"}:
        raise HTTPException(
            status_code=400,
            detail="student_selected_label must be positive, negative, or none."
        )

    exercise = (
        db.query(Exercise)
        .filter(
            Exercise.id == payload.exercise_id,
            Exercise.user_id == current_user.id
        )
        .first()
    )

    if not exercise:
        raise HTTPException(status_code=404, detail="Exercise not found.")

    existing_submission = (
        db.query(Submission)
        .filter(
            Submission.exercise_id == exercise.id,
            Submission.user_id == current_user.id
        )
        .first()
    )

    if existing_submission:
        raise HTTPException(status_code=400, detail="This exercise has already been submitted.")

    computed_label = exercise.relationship_label or classify_correlation(exercise.pearson_r)
    is_correct = selected_label == computed_label

    submission = Submission(
        exercise_id=exercise.id,
        user_id=current_user.id,
        student_selected_label=selected_label,
        student_explanation=payload.student_explanation.strip(),
        student_pearson_r=payload.student_pearson_r,
        computed_pearson_r=exercise.pearson_r,
        computed_relationship_label=computed_label,
        is_correct_label=is_correct
    )

    db.add(submission)
    db.commit()
    db.refresh(submission)

    return {
        "message": "Exercise submitted successfully",
        "submission_id": submission.id,
        "exercise_id": submission.exercise_id,
        "student_selected_label": submission.student_selected_label,
        "computed_relationship_label": submission.computed_relationship_label,
        "computed_pearson_r": round(submission.computed_pearson_r, 6),
        "is_correct_label": submission.is_correct_label
    }


# my submissions
@router.get("/my-submissions")
def get_my_submissions(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    submissions = (
        db.query(Submission, Exercise)
        .join(Exercise, Submission.exercise_id == Exercise.id)
        .filter(Submission.user_id == current_user.id)
        .order_by(Submission.submitted_at.desc())
        .all()
    )

    results = []
    for submission, exercise in submissions:
        results.append({
            "id": submission.id,
            "exercise_id": submission.exercise_id,
            "indicator_1_code": exercise.indicator_1_code,
            "indicator_1_name": exercise.indicator_1_name,
            "indicator_2_code": exercise.indicator_2_code,
            "indicator_2_name": exercise.indicator_2_name,
            "student_selected_label": submission.student_selected_label,
            "student_explanation": submission.student_explanation,
            "student_pearson_r": submission.student_pearson_r,
            "computed_relationship_label": submission.computed_relationship_label,
            "computed_pearson_r": submission.computed_pearson_r,
            "is_correct_label": submission.is_correct_label,
            "teacher_feedback": submission.teacher_feedback,
            "teacher_grade": submission.teacher_grade,
            "submitted_at": submission.submitted_at
        })

    return results


@router.get("/my-submissions/{submission_id}")
def get_one_my_submission(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    row = (
        db.query(Submission, Exercise)
        .join(Exercise, Submission.exercise_id == Exercise.id)
        .filter(
            Submission.id == submission_id,
            Submission.user_id == current_user.id
        )
        .first()
    )

    if not row:
        raise HTTPException(status_code=404, detail="Submission not found.")

    submission, exercise = row

    return {
        "id": submission.id,
        "exercise_id": submission.exercise_id,
        "indicator_1_code": exercise.indicator_1_code,
        "indicator_1_name": exercise.indicator_1_name,
        "indicator_2_code": exercise.indicator_2_code,
        "indicator_2_name": exercise.indicator_2_name,
        "student_selected_label": submission.student_selected_label,
        "student_explanation": submission.student_explanation,
        "computed_relationship_label": submission.computed_relationship_label,
        "computed_pearson_r": submission.computed_pearson_r,
        "is_correct_label": submission.is_correct_label,
        "teacher_feedback": submission.teacher_feedback,
            "teacher_grade": submission.teacher_grade,
        "submitted_at": submission.submitted_at
    }


# teacher submissions
@router.get("/submissions")
def get_all_submissions_for_teacher(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    require_teacher(current_user)

    rows = (
        db.query(Submission, Exercise, User)
        .join(Exercise, Submission.exercise_id == Exercise.id)
        .join(User, Submission.user_id == User.id)
        .order_by(Submission.submitted_at.desc())
        .all()
    )

    results = []
    for submission, exercise, user in rows:
        results.append({
            "id": submission.id,
            "exercise_id": submission.exercise_id,
            "student_id": user.id,
            "student_name": user.first_name,
            "student_email": user.email,
            "indicator_1_code": exercise.indicator_1_code,
            "indicator_1_name": exercise.indicator_1_name,
            "indicator_2_code": exercise.indicator_2_code,
            "indicator_2_name": exercise.indicator_2_name,
            "student_selected_label": submission.student_selected_label,
            "student_explanation": submission.student_explanation,
            "student_pearson_r": submission.student_pearson_r,
            "computed_relationship_label": submission.computed_relationship_label,
            "computed_pearson_r": submission.computed_pearson_r,
            "is_correct_label": submission.is_correct_label,
            "teacher_feedback": submission.teacher_feedback,
            "teacher_grade": submission.teacher_grade,
			"teacher_grade": submission.teacher_grade,
            "submitted_at": submission.submitted_at
        })

    return results


@router.get("/submissions/{submission_id}")
def get_one_submission_for_teacher(
    submission_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    require_teacher(current_user)

    row = (
        db.query(Submission, Exercise, User)
        .join(Exercise, Submission.exercise_id == Exercise.id)
        .join(User, Submission.user_id == User.id)
        .filter(Submission.id == submission_id)
        .first()
    )

    if not row:
        raise HTTPException(status_code=404, detail="Submission not found.")

    submission, exercise, user = row

    return {
        "id": submission.id,
        "exercise_id": submission.exercise_id,
        "student_id": user.id,
        "student_name": user.first_name,
        "student_email": user.email,
        "indicator_1_code": exercise.indicator_1_code,
        "indicator_1_name": exercise.indicator_1_name,
        "indicator_2_code": exercise.indicator_2_code,
        "indicator_2_name": exercise.indicator_2_name,
        "countries": json.loads(exercise.countries_json),
        "start_year": exercise.start_year,
        "end_year": exercise.end_year,
        "rows": json.loads(exercise.merged_rows_json),
        "student_selected_label": submission.student_selected_label,
        "student_explanation": submission.student_explanation,
        "computed_relationship_label": submission.computed_relationship_label,
        "computed_pearson_r": submission.computed_pearson_r,
        "is_correct_label": submission.is_correct_label,
        "teacher_feedback": submission.teacher_feedback,
            "teacher_grade": submission.teacher_grade,
        "teacher_grade": submission.teacher_grade,
        "submitted_at": submission.submitted_at
    }


# feedback
@router.patch("/submissions/{submission_id}/feedback")
def add_teacher_feedback(
    submission_id: int,
    payload: SubmissionFeedbackRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    require_teacher(current_user)

    submission = db.query(Submission).filter(Submission.id == submission_id).first()

    if not submission:
        raise HTTPException(status_code=404, detail="Submission not found.")

    submission.teacher_feedback = payload.teacher_feedback.strip()
    if payload.teacher_grade is not None:
        submission.teacher_grade = payload.teacher_grade.strip()
    db.commit()
    db.refresh(submission)

    return {
        "message": "Teacher feedback saved successfully",
        "submission_id": submission.id,
        "teacher_feedback": submission.teacher_feedback,
            "teacher_grade": submission.teacher_grade,
        "teacher_grade": submission.teacher_grade
    }






# data
@router.get("/data")
def get_data(
    indicator_code: str = Query(...),
    country_code: str | None = Query(None),
    limit: int = Query(100),
    db: Session = Depends(get_db)
):
    query = db.query(GapminderData).filter(
        GapminderData.indicator_code == indicator_code
    )

    if country_code:
        query = query.filter(GapminderData.country_code == country_code)

    rows = query.order_by(GapminderData.year.asc()).limit(limit).all()

    return [
        {
            "country": r.country if r.country else r.country_code,
            "country_code": r.country_code,
            "year": r.year,
            "value": r.value,
            "indicator_code": r.indicator_code
        }
        for r in rows
    ]


# countries
@router.get("/countries")
def get_countries(
    indicator_code: str = Query(...),
    year: int | None = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(
        GapminderData.country_code,
        GapminderData.country
    ).filter(
        GapminderData.indicator_code == indicator_code
    )

    if year is not None:
        query = query.filter(GapminderData.year == year)

    rows = (
        query.distinct()
        .order_by(GapminderData.country_code.asc())
        .all()
    )

    return [
        {
            "country_code": row.country_code,
            "country": row.country if row.country else row.country_code
        }
        for row in rows
    ]


# years
@router.get("/years")
def get_years(
    indicator_code: str = Query(...),
    db: Session = Depends(get_db)
):
    rows = (
        db.query(GapminderData.year)
        .filter(GapminderData.indicator_code == indicator_code)
        .distinct()
        .order_by(GapminderData.year.asc())
        .all()
    )

    return [row[0] for row in rows]


@router.get("/compare")
def compare_countries(
    indicator_code: str = Query(...),
    year: int = Query(...),
    country_codes: str = Query(..., description="Comma-separated list like abw,usa,can"),
    db: Session = Depends(get_db)
):
    codes = [code.strip().lower() for code in country_codes.split(",") if code.strip()]

    rows = (
        db.query(GapminderData)
        .filter(
            GapminderData.indicator_code == indicator_code,
            GapminderData.year == year,
            GapminderData.country_code.in_(codes)
        )
        .order_by(GapminderData.country_code.asc())
        .all()
    )

    return [
        {
            "country": r.country if r.country else r.country_code,
            "country_code": r.country_code,
            "year": r.year,
            "value": r.value,
            "indicator_code": r.indicator_code
        }
        for r in rows
    ]


# ── Students management (teacher only) ────────────────────────────────────────

@router.get("/students")
def get_students(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    require_teacher(current_user)
    from sqlalchemy import text
    students = db.query(User).filter(User.role == "student").order_by(User.section.asc(), User.first_name.asc()).all()
    return [
        {
            "id": s.id,
            "first_name": s.first_name,
            "email": s.email,
            "section": s.section or "Unassigned",
            "submission_count": db.query(Submission).filter(Submission.user_id == s.id).count(),
            "exercise_count": db.query(Exercise).filter(Exercise.user_id == s.id).count()
        }
        for s in students
    ]


@router.delete("/students/{user_id}")
def delete_student(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    require_teacher(current_user)

    student = db.query(User).filter(User.id == user_id, User.role == "student").first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found.")

    # Delete all submissions and exercises
    submission_ids = [s.id for s in db.query(Submission).filter(Submission.user_id == user_id).all()]
    db.query(Submission).filter(Submission.user_id == user_id).delete()
    db.query(Exercise).filter(Exercise.user_id == user_id).delete()
    db.delete(student)
    db.commit()

    return {"message": f"Student {student.first_name} and all their data deleted."}


# ── Announcements ──────────────────────────────────────────────────────────────

@router.get("/announcements")
def get_announcements(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.announcement import Announcement
    query = db.query(Announcement)

    # Students only see announcements for their section or all
    if current_user.role not in ("teacher", "instructor"):
        user_section = getattr(current_user, "section", None)
        query = query.filter(
            (Announcement.target_section == None) |
            (Announcement.target_section == user_section)
        )

    announcements = query.order_by(Announcement.created_at.desc()).all()
    return [
        {
            "id": a.id,
            "teacher_name": a.teacher_name,
            "message": a.message,
            "target_section": a.target_section,
            "created_at": a.created_at
        }
        for a in announcements
    ]


@router.post("/announcements")
def create_announcement(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.announcement import Announcement
    require_teacher(current_user)

    message = payload.get("message", "").strip()
    target_section = payload.get("target_section") or None

    if not message:
        raise HTTPException(status_code=400, detail="Message cannot be empty.")

    ann = Announcement(
        teacher_id=current_user.id,
        teacher_name=current_user.first_name,
        message=message,
        target_section=target_section
    )
    db.add(ann)
    db.commit()
    db.refresh(ann)

    return {
        "id": ann.id,
        "teacher_name": ann.teacher_name,
        "message": ann.message,
        "target_section": ann.target_section,
        "created_at": ann.created_at
    }


@router.delete("/announcements/{announcement_id}")
def delete_announcement(
    announcement_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    from app.models.announcement import Announcement
    require_teacher(current_user)

    ann = db.query(Announcement).filter(Announcement.id == announcement_id).first()
    if not ann:
        raise HTTPException(status_code=404, detail="Announcement not found.")

    db.delete(ann)
    db.commit()
    return {"message": "Announcement deleted."}